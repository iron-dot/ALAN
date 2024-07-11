import time
import threading
import pyttsx3
import speech_recognition as sr
from gtts import gTTS
from playsound import playsound
from openai import OpenAI  # OpenAI official Python package
import openpyxl
from openpyxl.styles import Alignment
import os
from datetime import datetime

# 음성 알림을 위한 초기화
engine = pyttsx3.init()

# 엑셀 파일 초기화 및 캘린더 생성 함수
def initialize_calendar(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendar"

    # 첫 번째 행에 월 이름 쓰기
    ws.merge_cells('A1:G1')
    ws['A1'] = "Calendar"
    ws['A1'].alignment = Alignment(horizontal="center")
    
    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for col, day in enumerate(days, 1):
        ws.cell(row=2, column=col, value=day)

    # 날짜 채우기 (임의로 6월 달력 예시)
    month_days = [
        ["", "", "", "", "", "", 1],
        [2, 3, 4, 5, 6, 7, 8],
        [9, 10, 11, 12, 13, 14, 15],
        [16, 17, 18, 19, 20, 21, 22],
        [23, 24, 25, 26, 27, 28, 29],
        [30, "", "", "", "", "", ""]
    ]

    start_row = 3
    for week in month_days:
        for col, day in enumerate(week, 1):
            ws.cell(row=start_row, column=col, value=day)
        start_row += 1

    wb.save(filename)

# 일정 엑셀 파일에 저장 함수
def save_schedule_to_excel(date, time, message):
    year_month = date[:7]  # YYYY-MM 형식으로 추출
    filename = f"{year_month}.xlsx"

    if not os.path.exists(filename):
        initialize_calendar(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # 일정 추가
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date)
    ws.cell(row=row, column=2, value=time)
    ws.cell(row=row, column=3, value=message)

    wb.save(filename)

# 일정 확인 함수
def check_schedules(schedules):
    while True:
        current_time = time.localtime()
        current_datetime = time.strftime("%Y-%m-%d %H:%M", current_time)

        for schedule_datetime, message in schedules.items():
            if current_datetime == schedule_datetime:
                
                print(f"알람 시간입니다! 스케줄: {message}")
                text = f"알람 시간입니다! 스케줄: {message}"
                tts = gTTS(text=text, lang='ko')
                import uuid

                filename = f"D:/InfiniteCore Industries/Alan/dataset/response_{uuid.uuid4()}.mp3"
                tts.save(filename)
                playsound(filename)

                # 음성 알림
                engine.say(message)
                engine.runAndWait()
                time.sleep(60)  # 1분 동안 대기 (중복 알림 방지)

        time.sleep(1)  # 1초 동안 대기

# 음성 명령 인식 함수
def recognize_voice_command():
    command = input("명령을 말하세요:")
    return command
    '''recognizer = sr.Recognizer()
    mic = sr.Microphone()

    with mic as source:
        print("명령을 말하세요:")
        audio = recognizer.listen(source)
        try:
            command = recognizer.recognize_google(audio, language='ko-KR')
            print(f"인식된 명령: {command}")
            return command
        except sr.UnknownValueError:
            print("입력을 인식하지 못했습니다. 다시 시도하세요.")
            return None
        except sr.RequestError as e:
            print(f"구글 음성 인식 서비스에 문제가 있습니다: {e}")
            return None'''

# 일정 확인 음성 출력 함수
def announce_schedule(date):
    year_month = date[:7]  # YYYY-MM 형식으로 추출
    filename = f"{year_month}.xlsx"

    if not os.path.exists(filename):
        response_text = "해당 날짜에 일정이 없습니다."
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        schedules = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == date:
                schedules.append(f"{row[1]}에 {row[2]}")

        if schedules:
            response_text = f"{date}의 일정은 다음과 같습니다: " + ", ".join(schedules)
        else:
            response_text = "해당 날짜에 일정이 없습니다."

    print(response_text)
    tts = gTTS(text=response_text, lang='ko')
    import uuid

    filename = f"D:/InfiniteCore Industries/Alan/dataset/response_{uuid.uuid4()}.mp3"
    tts.save(filename)
    playsound(filename)
# 음성 인식 스케줄 추가 함수
def add_schedule_by_voice(schedule_input):
    try:
        client = OpenAI(
        api_key="   ")
        # 상대적인 날짜 처리
        

        current_datetime = datetime.now()
        weekday = current_datetime.strftime("%A")
        weekday_korean = {
            'Monday': '월요일',
            'Tuesday': '화요일',
            'Wednesday': '수요일',
            'Thursday': '목요일',
            'Friday': '금요일',
            'Saturday': '토요일',
            'Sunday': '일요일'
        }
        formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M")
        print("현재 날짜 및 시간:", formatted_datetime)
        print("오늘의 요일:", weekday_korean[weekday])
        today = formatted_datetime
        weekday = weekday_korean[weekday]
        gpt_prompt = [{"role": "system", "content":
        f"""
        현재 지금 날짜와 시간 : {today}와 {weekday}를 바탕으로 
       사용자가 말한 날짜, 시간과 내용을 분리하여 날짜를 YYYY-MM-DD 형식으로, 시간을 24시간 형식의 HH:MM으로 변환하고, 내용을 추출해 주세요.
                예시 입력: "오늘 오전 9시에 회의"
                예시 출력: 2024-06-10 09:00, 회의
        입력: "{schedule_input}"
        출력:
        """}]


        response = client.chat.completions.create(
            model="gpt-4-1106-preview",
            messages=gpt_prompt,

        )

        input_prompt = response.choices[0].message.content
        print(input_prompt)
        schedule_time, schedule_message = input_prompt.split(',')

        print(f"변환된 시간: {schedule_time}")
        print(f"변환된 내용: {schedule_message}")

        date, time = schedule_time.split(' ')
        schedules[f"{date} {time}"] = schedule_message
        save_schedule_to_excel(date, time, schedule_message)
    except Exception as e:
        print(f"시간 변환 중 오류가 발생했습니다: {e}")

# 명령 처리 함수
def process_command(command):
    try:
        client = OpenAI(
        api_key="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")
        gpt_prompt = [{"role": "system", "content": """
                사용자가 말한 명령을 분석하여 '추가' 또는 '확인'인지 파악해 주세요.
                예시 입력: "오늘 오전 9시에 회의 추가해줘"
                예시 출력: 추가

                예시 입력: "오늘 일정 확인해줘"
                예시 출력: 확인
                """},
                {"role": "user", "content": command}
            ]
        

                

        response = client.chat.completions.create(
            model="gpt-4-1106-preview",
            messages=gpt_prompt,

        )

        input_prompt = response.choices[0].message.content
        print(input_prompt)
        intent = input_prompt
        if intent == "추가":
            add_schedule_by_voice(command)

        elif intent == "확인":

            # 상대적인 날짜 처리
            

            current_datetime = datetime.now()
            weekday = current_datetime.strftime("%A")
            weekday_korean = {
                'Monday': '월요일',
                'Tuesday': '화요일',
                'Wednesday': '수요일',
                'Thursday': '목요일',
                'Friday': '금요일',
                'Saturday': '토요일',
                'Sunday': '일요일'
            }
            formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M")
            print("현재 날짜 및 시간:", formatted_datetime)
            print("오늘의 요일:", weekday_korean[weekday])
            today = formatted_datetime
            weekday = weekday_korean[weekday]
            gpt_prompt = [
                {"role": "system", "content": f"""
                 항상 현재 날짜와 시간 : {today}와 {weekday}를 바탕으로 
                사용자가 말한 날짜를 YYYY-MM-DD 형식으로 변환해 주세요.
                예시 입력: "오늘 일정 확인해줘" 
                예시 출력: 2024-06-10
                 
                """},
                {"role": "user", "content": command}
            ]
            response = client.chat.completions.create(
            model="gpt-4-1106-preview",
            messages=gpt_prompt,

            )

            input_prompt = response.choices[0].message.content
            print(input_prompt)
            date = input_prompt
            
            
            announce_schedule(date)
        else:
            response_text = "알 수 없는 명령입니다. 다시 시도해 주세요."
            print(response_text)
            tts = gTTS(text=response_text, lang='ko')
            import uuid

            filename = f"D:/InfiniteCore Industries/Alan/dataset/response_{uuid.uuid4()}.mp3"
            tts.save(filename)
            playsound(filename)
    except Exception as e:
        print(f"명령 처리 중 오류가 발생했습니다: {e}")

if __name__ == "__main__":
    schedules = {}

    # 스케줄 확인 스레드 시작
    schedule_thread = threading.Thread(target=check_schedules, args=(schedules,))
    schedule_thread.daemon = True
    schedule_thread.start()

    # 음성 인식 명령 루프
    while True:
        command = ''#recognize_voice_command()
        if command:
            process_command(command)






'''import speech_recognition as sr

# 음성으로 일정 삭제 함수
def delete_schedule_by_voice():
    recognizer = sr.Recognizer()
    mic = sr.Microphone()

    with mic as source:
        print("삭제할 스케줄을 말하세요:")
        audio = recognizer.listen(source)
        try:
            schedule_to_delete = recognizer.recognize_google(audio, language='ko-KR')
            print(f"인식된 입력: {schedule_to_delete}")
        except sr.UnknownValueError:
            print("입력을 인식하지 못했습니다. 다시 시도하세요.")
            return
        except sr.RequestError as e:
            print(f"구글 음성 인식 서비스에 문제가 있습니다: {e}")
            return

    # 스케줄에서 삭제할 일정 찾기
    for schedule_datetime, schedule_message in schedules.items():
        if schedule_to_delete in schedule_message:
            del schedules[schedule_datetime]
            update_schedule_list()
            print(f"일정 '{schedule_message}'이 삭제되었습니다.")
            return

    print("해당하는 일정을 찾지 못했습니다.")

# GUI에 음성으로 일정 삭제 버튼 추가 및 이벤트 핸들러 등록 생략
'''